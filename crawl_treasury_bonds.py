#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
国债业务公告数据抓取脚本 v2
功能：
1. 抓取所有国债业务公告
2. 提取普通发行公告的关键数据
3. 推断续发行公告的期限（从偿还本金日期计算）
4. 拆分储蓄国债多债合并公告
5. 提取做市操作公告的表格数据
6. 生成多sheet的Excel文件
"""

import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from urllib.parse import urljoin
import time
from datetime import datetime

# 基础配置
BASE_URL = 'https://zwgls.mof.gov.cn/ywgg/'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

def get_list_page(page_num):
    """获取列表页内容"""
    if page_num == 1:
        url = BASE_URL
    else:
        url = urljoin(BASE_URL, f'index_{page_num-1}.htm')
    
    try:
        response = requests.get(url, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        return response.text
    except Exception as e:
        print(f"  获取列表页失败: {e}")
        return None

def extract_announcements_from_list(html):
    """从列表页提取公告链接"""
    soup = BeautifulSoup(html, 'html.parser')
    announcements = []
    
    # 查找所有链接
    for a_tag in soup.find_all('a', href=True):
        title = a_tag.get_text(strip=True)
        href = a_tag['href']
        
        # 只保留国债业务公告
        if '国债业务公告' in title and '第' in title and '号' in title:
            # 构建完整URL
            full_url = urljoin(BASE_URL, href)
            # 提取年份和编号
            year_match = re.search(r'(\d{4})年第(\d+)号', title)
            if year_match:
                year = int(year_match.group(1))
                num = int(year_match.group(2))
                announcements.append({
                    'title': title,
                    'url': full_url,
                    'year': year,
                    'num': num
                })
    
    return announcements

def get_detail_page(url):
    """获取详情页内容"""
    try:
        response = requests.get(url, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        return response.text
    except Exception as e:
        print(f"    获取详情页失败: {e}")
        return None

def is_market_operation_announcement(content):
    """判断是否是做市支持操作公告"""
    return '国债做市支持操作' in content or '做市支持操作' in content

def is_savings_bond_announcement(content):
    """判断是否是储蓄国债公告（可能包含多期）"""
    return '储蓄国债' in content and ('第三期' in content or '第四期' in content or '第二期' in content or '第一期' in content)

def is_reopening_announcement(content):
    """判断是否是续发行公告"""
    return '续发行' in content

def extract_market_operation_data(html, title, year, num, url):
    """提取做市操作公告的数据（从HTML表格中提取）"""
    results = []
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # 提取文件日期
    content = soup.get_text()
    file_date = ''
    match = re.search(r'中华人民共和国财政部\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
    if match:
        file_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    else:
        match = re.search(r'发布日期[:：]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
        if match:
            file_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    
    # 提取操作方向
    direction = ''
    if '随卖' in content:
        direction = '随卖'
    elif '随买' in content:
        direction = '随买'
    
    # 查找表格
    tables = soup.find_all('table')
    
    for table in tables:
        rows = table.find_all('tr')
        if len(rows) < 2:
            continue
        
        # 解析表头
        headers = []
        header_row = rows[0]
        for th in header_row.find_all(['th', 'td']):
            headers.append(th.get_text(strip=True))
        
        # 记录上一行的操作方向（处理合并单元格的情况）
        last_direction = direction
        
        # 解析数据行
        for row in rows[1:]:
            cells = row.find_all(['td', 'th'])
            if len(cells) < 3:
                continue
            
            row_data = [cell.get_text(strip=True) for cell in cells]
            
            # 尝试识别各列
            bond_name = ''
            period = ''
            amount = ''
            price = ''
            rate = ''
            current_direction = ''
            
            # 根据表头匹配
            for i, header in enumerate(headers):
                if i >= len(row_data):
                    break
                val = row_data[i]
                if '操作方向' in header:
                    current_direction = val
                elif '操作券种' in header or '券种' in header:
                    bond_name = val
                elif '期限' in header:
                    period = val + '年' if val and '年' not in val else val
                elif '操作额' in header or '金额' in header:
                    amount = val
                elif '中标价格' in header or '价格' in header:
                    price = val
                elif '收益率' in header:
                    rate = val + '%' if val and '%' not in val else val
            
            # 如果操作方向为空，使用上一行的操作方向
            if not current_direction or '国债' in current_direction:
                # 如果current_direction包含"国债"，说明匹配错了（第一列合并了）
                current_direction = last_direction
            else:
                last_direction = current_direction
            
            # 如果bond_name不包含"国债"，说明匹配错了，需要调整
            if bond_name and '国债' not in bond_name:
                # 可能是因为第一列合并了，所有数据左移了一列
                bond_name = ''
                # 重新按位置匹配（从第一列开始就是券种）
                if len(row_data) >= 1:
                    bond_name = row_data[0]
                if len(row_data) >= 2:
                    period = row_data[1]
                    if period and '年' not in period:
                        period += '年'
                if len(row_data) >= 3:
                    amount = row_data[2]
                if len(row_data) >= 4:
                    price = row_data[3]
                if len(row_data) >= 5:
                    rate = row_data[4]
                    if rate and '%' not in rate:
                        rate += '%'
            
            # 如果没匹配到表头，尝试按位置匹配
            if not bond_name:
                # 第一列可能是操作方向，第二列是券种
                if len(row_data) >= 1:
                    if row_data[0] in ('随卖', '随买'):
                        current_direction = row_data[0]
                        if len(row_data) >= 2:
                            bond_name = row_data[1]
                        if len(row_data) >= 3:
                            period = row_data[2]
                            if period and '年' not in period:
                                period += '年'
                        if len(row_data) >= 4:
                            amount = row_data[3]
                        if len(row_data) >= 5:
                            price = row_data[4]
                        if len(row_data) >= 6:
                            rate = row_data[5]
                            if rate and '%' not in rate:
                                rate += '%'
                    else:
                        # 第一列就是券种（操作方向列合并了）
                        bond_name = row_data[0]
                        if len(row_data) >= 2:
                            period = row_data[1]
                            if period and '年' not in period:
                                period += '年'
                        if len(row_data) >= 3:
                            amount = row_data[2]
                        if len(row_data) >= 4:
                            price = row_data[3]
                        if len(row_data) >= 5:
                            rate = row_data[4]
                            if rate and '%' not in rate:
                                rate += '%'
            
            # 只保留有效的数据行（有券种名称）
            if bond_name and '国债' in bond_name:
                result = {
                    '公告标题': title,
                    '年份': year,
                    '公告编号': num,
                    '国债类型': '附息国债',
                    '操作方向': current_direction,
                    '操作券种': bond_name,
                    '期限': period,
                    '操作额(亿元)': amount,
                    '中标价格(元)': price,
                    '收益率(%)': rate,
                    '文件日期': file_date,
                    '公告链接': url
                }
                results.append(result)
    
    return results

def extract_savings_bond_data(content, title, year, num, url):
    """提取储蓄国债数据（支持多期拆分）"""
    results = []
    
    # 提取文件日期
    file_date = ''
    match = re.search(r'中华人民共和国财政部\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
    if match:
        file_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    else:
        match = re.search(r'发布日期[:：]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
        if match:
            file_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    
    # 提取开始计息日期
    start_date = ''
    match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日起息', content)
    if match:
        start_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    else:
        match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日开始计息', content)
        if match:
            start_date = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    
    # 提取发行期
    issue_period = ''
    match = re.search(r'发行期为(\d{4})年(\d{1,2})月(\d{1,2})日至(\d{1,2})月(\d{1,2})日', content)
    if match:
        issue_period = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日至{match.group(4)}月{match.group(5)}日'
    
    # 查找各期国债信息
    # 模式：第X期期限X年，票面年利率为X%，最大发行额X亿元
    pattern = r'第([一二三四五六七八九十]+)期期限(\d+)年，票面年利率为([\d.]+)%，最大发行额([\d.]+)亿元'
    matches = re.findall(pattern, content)
    
    period_map = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
    
    for match in matches:
        period_name = f'第{match[0]}期'
        period_num = period_map.get(match[0], match[0])
        period = f'{match[1]}年'
        rate = f'{match[2]}%'
        amount = match[3]
        
        result = {
            '公告标题': title,
            '年份': year,
            '公告编号': num,
            '国债类型': f'储蓄国债（{period_name}）',
            '计划发行(亿元)': amount,
            '实际发行(亿元)': 'NA',
            '期限': period,
            '发行价格(元)': 'NA',
            '票面利率/收益率': rate,
            '文件日期': file_date,
            '开始计息': start_date,
            '上市交易/发行期': issue_period,
            '公告链接': url
        }
        results.append(result)
    
    return results

def extract_normal_bond_data(content, title, year, num, url):
    """提取普通国债/续发行国债的数据"""
    result = {
        '公告标题': title,
        '年份': year,
        '公告编号': num,
        '国债类型': '未知',
        '计划发行(亿元)': 'NA',
        '实际发行(亿元)': 'NA',
        '期限': 'NA',
        '发行价格(元)': 'NA',
        '票面利率/收益率': 'NA',
        '文件日期': 'NA',
        '开始计息': 'NA',
        '上市交易': 'NA',
        '公告链接': url
    }
    
    # 判断国债类型
    if '附息' in content:
        result['国债类型'] = '附息国债'
    elif '贴现' in content:
        result['国债类型'] = '贴现国债'
    elif '储蓄' in content:
        result['国债类型'] = '储蓄国债'
    elif '特别国债' in content:
        result['国债类型'] = '特别国债'
    
    # 判断是否是续发行
    is_reopening = is_reopening_announcement(content)
    if is_reopening:
        result['国债类型'] = result['国债类型'] + '（续发行）'
    
    # 计划发行金额（支持"计划发行"和"计划续发行"）
    match = re.search(r'计划[续]?发行([\d.]+)亿元', content)
    if match:
        result['计划发行(亿元)'] = match.group(1)
    
    # 实际发行面值金额（支持"实际发行"和"实际续发行"）
    match = re.search(r'实际[续]?发行面值金额([\d.]+)亿元', content)
    if match:
        result['实际发行(亿元)'] = match.group(1)
    else:
        match = re.search(r'实际[续]?发行([\d.]+)亿元', content)
        if match:
            result['实际发行(亿元)'] = match.group(1)
    
    # 期限提取
    # 1. 直接提取"X年"期限
    match = re.search(r'期限(\d+)年', content)
    if match:
        result['期限'] = match.group(1) + '年'
    else:
        # 2. 提取"X年期"
        match = re.search(r'(\d+)年期', content)
        if match:
            result['期限'] = match.group(1) + '年'
        else:
            # 3. 贴现国债可能是"X天"
            match = re.search(r'期限(\d+)天', content)
            if match:
                result['期限'] = match.group(1) + '天'
            else:
                # 4. 续发行公告：从偿还本金日期推断期限
                if is_reopening:
                    match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日偿还本金', content)
                    if match:
                        repay_year = int(match.group(1))
                        # 从国债名称中提取发行年份
                        name_match = re.search(r'(\d{4})年记账式', content)
                        if name_match:
                            issue_year = int(name_match.group(1))
                            period = repay_year - issue_year
                            if period > 0:
                                result['期限'] = f'{period}年'
    
    # 发行价格
    match = re.search(r'发行价格为([\d.]+)元', content)
    if match:
        result['发行价格(元)'] = match.group(1)
    else:
        # 续发行价格
        match = re.search(r'续发行价格为([\d.]+)元', content)
        if match:
            result['发行价格(元)'] = match.group(1)
        else:
            # 检查是否按面值发行
            if '按面值发行' in content:
                result['发行价格(元)'] = '100'
    
    # 票面利率/收益率
    # 1. 票面利率为X%
    match = re.search(r'票面利率为([\d.]+)%', content)
    if match:
        result['票面利率/收益率'] = match.group(1) + '%'
    else:
        # 2. 票面利率X%（没有"为"字的情况）
        match = re.search(r'票面利率([\d.]+)%', content)
        if match:
            result['票面利率/收益率'] = match.group(1) + '%'
        else:
            # 3. 折合年收益率（贴现国债）
            match = re.search(r'折合年收益率为([\d.]+)%', content)
            if match:
                result['票面利率/收益率'] = match.group(1) + '%'
            else:
                match = re.search(r'折合年收益率([\d.]+)%', content)
                if match:
                    result['票面利率/收益率'] = match.group(1) + '%'
                else:
                    # 4. 年收益率
                    match = re.search(r'年收益率为([\d.]+)%', content)
                    if match:
                        result['票面利率/收益率'] = match.group(1) + '%'
                    else:
                        match = re.search(r'年收益率([\d.]+)%', content)
                        if match:
                            result['票面利率/收益率'] = match.group(1) + '%'
                        else:
                            # 5. 票面年利率（储蓄国债）
                            match = re.search(r'票面年利率为([\d.]+)%', content)
                            if match:
                                result['票面利率/收益率'] = match.group(1) + '%'
    
    # 文件日期
    match = re.search(r'中华人民共和国财政部\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
    if match:
        result['文件日期'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    else:
        match = re.search(r'发布日期[:：]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', content)
        if match:
            result['文件日期'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    
    # 开始计息日期
    match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日开始计息', content)
    if match:
        result['开始计息'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    else:
        match = re.search(r'自(\d{4})年(\d{1,2})月(\d{1,2})日起计息', content)
        if match:
            result['开始计息'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
        else:
            match = re.search(r'起息日为(\d{4})年(\d{1,2})月(\d{1,2})日', content)
            if match:
                result['开始计息'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
    
    # 上市交易日期
    match = re.search(r'(\d{1,2})月(\d{1,2})日起上市交易', content)
    if match:
        result['上市交易'] = f'{match.group(1)}月{match.group(2)}日'
    else:
        match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日起上市交易', content)
        if match:
            result['上市交易'] = f'{match.group(1)}年{match.group(2)}月{match.group(3)}日'
        else:
            # 续发行的上市交易
            match = re.search(r'从(\d{1,2})月(\d{1,2})日起.*?上市交易', content)
            if match:
                result['上市交易'] = f'{match.group(1)}月{match.group(2)}日'
    
    return result

def extract_all_data():
    """提取所有公告数据"""
    all_announcements = []
    
    # 遍历所有列表页
    page_num = 1
    empty_pages = 0
    max_empty = 2
    
    print('开始抓取国债业务公告数据...')
    
    while empty_pages < max_empty:
        print(f'正在处理第 {page_num} 页...')
        html = get_list_page(page_num)
        
        if not html:
            empty_pages += 1
            page_num += 1
            continue
        
        announcements = extract_announcements_from_list(html)
        print(f'  找到 {len(announcements)} 条国债业务公告')
        
        if len(announcements) == 0:
            empty_pages += 1
        else:
            empty_pages = 0
            all_announcements.extend(announcements)
        
        page_num += 1
        time.sleep(0.5)  # 礼貌爬取
    
    print(f'\n共找到 {len(all_announcements)} 条国债业务公告')
    
    # 去重
    seen = set()
    unique_announcements = []
    for ann in all_announcements:
        key = (ann['year'], ann['num'])
        if key not in seen:
            seen.add(key)
            unique_announcements.append(ann)
    
    print(f'去重后共 {len(unique_announcements)} 条公告')
    
    # 按年份+编号降序排序（最新在前）
    unique_announcements.sort(key=lambda x: (x['year'], x['num']), reverse=True)
    
    # 抓取详情页数据
    print('\n开始抓取详情页数据...')
    
    normal_bonds = []      # 普通国债 + 续发行国债
    savings_bonds = []     # 储蓄国债（多期拆分）
    market_operations = [] # 做市操作
    
    for i, ann in enumerate(unique_announcements):
        print(f'  处理 {i+1}/{len(unique_announcements)}: {ann["title"]}')
        
        html = get_detail_page(ann['url'])
        if not html:
            continue
        
        soup = BeautifulSoup(html, 'html.parser')
        content = soup.get_text()
        
        # 判断公告类型，分别处理
        if is_market_operation_announcement(content):
            # 做市操作公告
            data = extract_market_operation_data(html, ann['title'], ann['year'], ann['num'], ann['url'])
            market_operations.extend(data)
        elif is_savings_bond_announcement(content):
            # 储蓄国债多期公告
            data = extract_savings_bond_data(content, ann['title'], ann['year'], ann['num'], ann['url'])
            savings_bonds.extend(data)
        else:
            # 普通国债/续发行国债
            data = extract_normal_bond_data(content, ann['title'], ann['year'], ann['num'], ann['url'])
            normal_bonds.append(data)
        
        time.sleep(0.3)  # 礼貌爬取
    
    return normal_bonds, savings_bonds, market_operations

def create_excel(normal_bonds, savings_bonds, market_operations, output_file):
    """生成多sheet的Excel文件"""
    wb = openpyxl.Workbook()
    
    # 定义样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: 国债发行数据（普通 + 续发行）
    ws1 = wb.active
    ws1.title = '国债发行数据'
    
    if normal_bonds:
        headers = list(normal_bonds[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row, data in enumerate(normal_bonds, 2):
            for col, header in enumerate(headers, 1):
                ws1.cell(row=row, column=col, value=data.get(header, ''))
    
    # 自适应列宽
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width
    
    # 冻结首行
    ws1.freeze_panes = 'A2'
    
    # Sheet 2: 储蓄国债数据
    ws2 = wb.create_sheet('储蓄国债数据')
    
    if savings_bonds:
        headers = list(savings_bonds[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row, data in enumerate(savings_bonds, 2):
            for col, header in enumerate(headers, 1):
                ws2.cell(row=row, column=col, value=data.get(header, ''))
    
    # 自适应列宽
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column].width = adjusted_width
    
    ws2.freeze_panes = 'A2'
    
    # Sheet 3: 做市操作数据
    ws3 = wb.create_sheet('做市操作数据')
    
    if market_operations:
        headers = list(market_operations[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row, data in enumerate(market_operations, 2):
            for col, header in enumerate(headers, 1):
                ws3.cell(row=row, column=col, value=data.get(header, ''))
    
    # 自适应列宽
    for col in ws3.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws3.column_dimensions[column].width = adjusted_width
    
    ws3.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_file)
    print(f'\n数据已保存到: {output_file}')

def main():
    output_file = '/Users/ankangli/.doubao/chats/2026-06-25/new-chat/国债业务公告数据_v2.xlsx'
    
    normal_bonds, savings_bonds, market_operations = extract_all_data()
    
    # 生成Excel
    create_excel(normal_bonds, savings_bonds, market_operations, output_file)
    
    # 打印统计信息
    print(f'\n统计:')
    print(f'  时间范围: {normal_bonds[-1]["年份"]}年第{normal_bonds[-1]["公告编号"]}号 - {normal_bonds[0]["年份"]}年第{normal_bonds[0]["公告编号"]}号' if normal_bonds else '')
    print(f'  国债发行数据: {len(normal_bonds)} 条')
    print(f'  储蓄国债数据: {len(savings_bonds)} 条')
    print(f'  做市操作数据: {len(market_operations)} 条')
    
    # 预览前几条
    print('\n国债发行数据预览（前5条）:')
    print(f'{'公告标题':<30} {'类型':<15} {'计划发行':<10} {'期限':<8} {'利率':<8}')
    print('-' * 75)
    for data in normal_bonds[:5]:
        print(f'{str(data["公告标题"])[:28]:<30} {str(data["国债类型"])[:13]:<15} {str(data["计划发行(亿元)"]):<10} {str(data["期限"]):<8} {str(data["票面利率/收益率"]):<8}')
    
    if savings_bonds:
        print('\n储蓄国债数据预览（前5条）:')
        print(f'{'公告标题':<30} {'类型':<15} {'计划发行':<10} {'期限':<8} {'利率':<8}')
        print('-' * 75)
        for data in savings_bonds[:5]:
            print(f'{str(data["公告标题"])[:28]:<30} {str(data["国债类型"])[:13]:<15} {str(data["计划发行(亿元)"]):<10} {str(data["期限"]):<8} {str(data["票面利率/收益率"]):<8}')
    
    if market_operations:
        print('\n做市操作数据预览（前5条）:')
        print(f'{'公告标题':<30} {'操作券种':<25} {'期限':<8} {'操作额':<8} {'收益率':<8}')
        print('-' * 85)
        for data in market_operations[:5]:
            print(f'{str(data["公告标题"])[:28]:<30} {str(data["操作券种"])[:23]:<25} {str(data["期限"]):<8} {str(data["操作额(亿元)"]):<8} {str(data["收益率(%)"]):<8}')

if __name__ == '__main__':
    main()
